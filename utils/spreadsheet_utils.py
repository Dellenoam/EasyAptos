import os
from typing import Dict, List, Tuple

from openpyxl import Workbook

from config.user_config import debug_mode


def read_wallets() -> List[Tuple[str, str, str]] | None:
    private_keys = []
    destination_addresses = []

    private_keys_filename = "private_keys.txt"
    destination_addresses_filename = "recipients.txt"

    if not os.path.exists(private_keys_filename) or not os.path.exists(
        destination_addresses_filename
    ):
        return None

    with open(private_keys_filename, "r") as f:
        private_keys = [line.strip() for line in f.readlines()]

    with open(destination_addresses_filename, "r") as f:
        destination_addresses = [line.strip() for line in f.readlines()]

    if len(private_keys) != len(destination_addresses):
        return None

    wallets = []

    for i, (private_key, destination_address) in enumerate(
        zip(private_keys, destination_addresses)
    ):
        name = str(i + 1)
        if not private_key.startswith("ed25519-priv-"):
            private_key = "ed25519-priv-" + private_key

        wallets.append((name, private_key, destination_address))

    return wallets


def write_balances_to_xlsx(
    filename: str,
    all_wallet_balances: List[Tuple[str, Dict[str, Dict[str, int]]]],
) -> None:
    workbook = Workbook()

    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    sheet = workbook.create_sheet(title="Balances")

    sheet.append(["Wallet Name", "Wallet Address", "APT"])

    for wallet_name, wallet_data in all_wallet_balances:
        for wallet_address, balances in wallet_data.items():
            apt_balance = balances.get("APT", "N/A")
            row = [wallet_name, wallet_address, apt_balance]
            sheet.append(row)

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    workbook.save(filename)


def write_transactions_to_xlsx(
    filename: str, all_transaction_data: List[Tuple[str, str, int]]
) -> None:
    workbook = Workbook()

    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    sheet = workbook.create_sheet(title="Transactions")

    sheet.append(["Wallet Name", "Transaction Link", "Transfered Amount (APT)"])

    for wallet_name, txn_hash, transfered_amount in all_transaction_data:
        if txn_hash != "N/A":
            aptos_explorer_transaction_url = "https://explorer.aptoslabs.com/txn/"
            aptos_explorer_transaction_url += txn_hash
            if debug_mode:
                aptos_explorer_transaction_url += "?network=testnet"
            else:
                aptos_explorer_transaction_url += "?network=mainnet"
        else:
            aptos_explorer_transaction_url = "N/A"

        sheet.append([wallet_name, aptos_explorer_transaction_url, transfered_amount])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        workbook.save(filename)


def write_create_wallets_to_xlsx(wallets: List[Tuple[str, str]], filename: str) -> None:
    workbook = Workbook()

    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    sheet = workbook.create_sheet(title="Wallets")

    sheet.append(["Name", "Wallet Address", "Wallet Private Key"])
    for index, (private_key, destination_address) in enumerate(wallets, start=1):
        name = str(index)
        sheet.append([name, private_key, destination_address])

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    workbook.save(filename)
